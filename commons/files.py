import yaml
import os
import pandas as pd
import openpyxl
from configparser import ConfigParser


class FileHandler:
    """处理文件读取和写入操作的类"""

    @staticmethod
    def read_file(path):
        """根据文件路径的后缀名读取文件内容"""
        _, file_extension = os.path.splitext(path)
        file_extension = file_extension.lower()

        try:
            if file_extension == '.yaml' or file_extension == '.yml':
                with open(path, 'r', encoding='utf-8') as fp:
                    # 返回dict类型
                    return yaml.safe_load(fp)
            elif file_extension == '.ini' or file_extension == '.conf':
                config = ConfigParser()
                with open(path, 'r', encoding='utf-8') as fp:
                    config.read_file(fp)
                    return config
            elif file_extension == '.csv':
                return pd.read_csv(path)
            elif file_extension == '.xlsx':
                return pd.read_excel(path, engine='openpyxl')
            else:
                with open(path, 'r', encoding='utf-8') as fp:
                    return fp.read()
        except Exception as e:
            print(f"Error reading file {path}: {e}")
            return None

    @staticmethod
    def write_file(path, data):
        """根据文件路径的后缀名写入文件内容"""
        _, file_extension = os.path.splitext(path)
        file_extension = file_extension.lower()

        try:
            if file_extension == '.yaml' or file_extension == '.yml':
                with open(path, 'w', encoding='utf-8') as fp:
                    yaml.dump(data, fp, default_flow_style=False)
            elif file_extension == '.ini' or file_extension == '.conf':
                config = ConfigParser()
                for section, options in data.items():
                    config[section] = dict(options)
                with open(path, 'w', encoding='utf-8') as fp:
                    config.write(fp)
            elif file_extension == '.csv':
                pd.DataFrame(data).to_csv(path, index=False, encoding='utf-8')
            elif file_extension == '.xlsx':
                FileHandler._write_excel(path, data)
            else:
                with open(path, 'w', encoding='utf-8') as fp:
                    fp.write(str(data))
        except Exception as e:
            print(f"Error writing to file {path}: {e}")

    @staticmethod
    def _write_excel(path, data):
        """将数据写入Excel文件"""
        workbook = openpyxl.Workbook()
        for sheet_name, sheet_data in data.items():
            worksheet = workbook.create_sheet(sheet_name)
            for row in range(len(sheet_data)):
                for col, value in enumerate(sheet_data[row]):
                    worksheet.cell(row=row + 1, column=col + 1).value = value
        workbook.save(path)


# 示例用法
file_handler = FileHandler()

# 读取文件示例
# yaml_data = file_handler.read_file('example.yaml')
# print(type(yaml_data))
# ini_data = file_handler.read_file('example.ini')
# print(ini_data[""][""])
# conf_data = file_handler.read_file('example.conf')
# print(conf_data["database"]["host"])
# csv_data = file_handler.read_file('example.csv')
# print(csv_data)
# excel_data = file_handler.read_file('example.xlsx')
# print(excel_data)
# 写入文件示例
# yaml_data_to_write = {'key': 'value'}
# file_handler.write_file('output.yaml', yaml_data_to_write, 'yaml')
#
# config_data_to_write = {'Section1': {'Option1': 'Value1'}}
# file_handler.write_file('output.ini', config_data_to_write, 'ini')
#
# csv_data_to_write = {'Column1': [1, 2, 3], 'Column2': ['a', 'b', 'c']}
# file_handler.write_file('output.csv', csv_data_to_write, 'csv')
#
# excel_data_to_write = {
#     'Sheet1': [[1, 'a'], [2, 'b'], [3, 'c']],
#     'Sheet2': [[4, 'd'], [5, 'e'], [6, 'f']]
# }
# file_handler.write_file('output.xlsx', excel_data_to_write, 'excel')
